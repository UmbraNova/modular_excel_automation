# from openpyxl import *
from openpyxl import Workbook, load_workbook



wb = Workbook()

# creating sheets
ws = wb.create_sheet("A sheet", 0)  # insert it in position 0, could add it to a variable
ws1 = wb.create_sheet("B sheet")  # if it's the same name the it will add 1 to the and
wb.create_sheet("C sheet")  # will append to the end if added without index

# removing sheets
wb.remove(wb["Sheet"])
del wb["B sheet"]

# renaming sheets
ws.title = "New title"

# showing sheets names
# for sheet in wb:
    # print(sheet)

wb.save("new sheets.xlsx")
# /////////////////////////////////

wb = load_workbook("new sheets.xlsx")

for sheet in wb:
    print(sheet.title)


# copy sheets
source = wb["New title"]
new_sheet = wb.copy_worksheet(source)

new_sheet.title = "Copied New title"
wb.save("copied_sheets.xlsx")

# acces using index
print("="*30)
ws2 = wb.worksheets[0]
print(ws2)
print(wb.sheetnames)

# access sheet that doesnt exist, gives an ERROR!
ws3 = wb["hello world!"]

# /////////////////////////////////






















