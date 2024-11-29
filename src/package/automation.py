import pandas as pd
import openpyxl
import os


# Filter the DataFrame to only include the desired columns


def process_file(file_path, output_ws, column_name_variants=["description", "denumire produs", "name"]):
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    header_seq = find_table_header(ws)

def find_table_header(ws):
    count_to_pass = 0
    # col_mapping = {"description": "", "gross weight": "", "net weight": ""}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                if in_required_columns(str(cell.value).lower()):
                    print(f"Letter: [{cell.column_letter}] | Value: [{cell.value}]")
                    # col_mapping["description"] = cell.column_letter
                    count_to_pass += 1
        if count_to_pass >= 4:  # if n num of matches found, means this is the table header
            break
        else:
            pass  # TODO: reset the col_mapping


def in_required_columns(cell_value):
    required_columns = {
        "sku_net_weight_col": ["greutate neta", "gerutate neta", "net", "neta", "netto"],
        "sku_description_col": ["denumire produs", "denumire produse", "description", "descriere"],
        "sku_gross_weight_col": ["greutate bruta", "gerutate bruta", "brutto", "brut", "gross"],
        "sku_length_col": ["lungime", "length"],
        "sku_width_col": ["latime", "width"],
        "sku_height_col": ["inalatime", "height"],
        "sku_bax_col": ["bax", "baxaj", "baxare", "qty/bax"],
        "sku_ean_col": ["ean", "cod ean", "barcod", "barcode"],
        "sku_quantity_col": ["cantitate", "quantity", "cantitate estimativa"],
        "sku_tariff_col": ["vamal", "hs-code", "hscode", "tariff no.", "tariff", "hs code"]
    }


    for values in required_columns.values():
        for i_val in values:
            cell_value = cell_value.replace(",", " ").replace(".", " ")
            if cell_value in i_val or i_val in cell_value:
                return True


def process_folders(base_folder, output_template):
    output_wb = openpyxl.load_workbook(output_template)
    output_ws = output_wb.active

    for root, dirs, files in os.walk(base_folder):
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(root, file)
                print(f"Processing file: {file}")
                process_file(file_path, output_ws)

    output_wb.save(output_template)
    print(f"All processed and saved in {output_template}")



base_folder = "X:/AUR/modular_excel_test/data"
output_template = "X:/AUR/modular_excel_test/result/output_file.xlsx"

process_folders(base_folder, output_template)
