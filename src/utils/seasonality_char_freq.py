import openpyxl



wb = openpyxl.load_workbook("X:/AUR/11.2024/25.11.2024+/sezonalitate!/!with season codes - Default25_11_2024_10_33_17 - Copy.xlsx")
ws = wb.active


# for row in ws.iter_cols():

desc_correlation_count = dict()


for row in ws.iter_rows():
    for cell in row[1:]:
        cell_value = str(cell.value)
        if cell_value.isdigit():
            print(f"Seasonality Code: {cell_value}")
            desc_correlation_count
        else:
            print(f"Description: {cell_value}")


















