# compara lista de produse to be imported cu produsele care deja sunt in system
# mereu de re-downloadat fisierul din configuration packages > ITEM > Excel > Export to Excel
# le coloreaza (articolul) cu albastru

# ulterior schimbat cu ALEX_BARCODES, comparare barcodes

# TODO: de facut automat si universal pentru sablonul de import

import openpyxl
from openpyxl.styles import PatternFill


def main(template_file_path="X:/AUR/11.2024/21.11.2024i/listing_new items+container 46/noul template - Copy.xlsx",
         items_in_system_path="X:/AUR/11.2024/21.11.2024i/listing_new items+container 46/barcodes_in_system.xlsx"):
    
    def compare_and_update(file_one_path, file_two_path, sheet_one_name='Item', sheet_two_name='Barcodes'):
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        workbook_one = openpyxl.load_workbook(file_one_path)
        sheet_one = workbook_one[sheet_one_name]
        workbook_two = openpyxl.load_workbook(file_two_path)
        sheet_two = workbook_two[sheet_two_name]
        
        # Collect values from file two's specified column into a set for faster lookup
        file_two_values = {sheet_two[f"A{row}"].value for row in range(1, sheet_two.max_row + 1) if sheet_two[f"A{row}"].value}

        for row_one in range(1, sheet_one.max_row + 1):
            try:
                cell_one = sheet_one[f"AF{row_one}"]
                cell_one_value = cell_one.value
                if str(cell_one_value) in str(file_two_values):
                    cell_one.fill = blue_fill
                    print(f"Match found: {cell_one_value} at row {row_one}, marked in blue.")
            except:
                break

        workbook_one.save(file_one_path)
        print("Workbook saved successfully!")

    compare_and_update(template_file_path, items_in_system_path)

if __name__ == "__main__":
    main()
