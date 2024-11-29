import openpyxl
import os
import zipfile

# given a folder with images, it writes the names in the created excel file


def process_folders(workbook_path, base_folder):
    output_wb = openpyxl.load_workbook(workbook_path)
    counted_row = 0

    if zipfile.is_zipfile(base_folder):
        with zipfile.ZipFile(base_folder, "r") as zip_ref:
            base_folder = base_folder[:-4]
            zip_ref.extractall(base_folder)
    for root,dirs, files in os.walk(base_folder):
        for file in files:
            if file.endswith(".jpg") or file.endswith(".jifi") or file.endswith(".png"):
                counted_row += 1
                write_in_excel(output_wb, file, counted_row)
    output_wb.save(workbook_path)


def write_in_excel(output_wb, file, counted_row):
    output_ws = output_wb.active
    output_ws.cell(row=counted_row, column=1).value = file[:-4]  # removes the last for chars

    # for row in output_ws.iter_rows():
    #     for cell in row:
    #         if cell.value:
    #             cell.value = file
    #             print(file)    
    


workbook_path = "X:/AUR/11.2024/25.11.2024/import poze/pictures_import.xlsx"
if not os.path.exists(workbook_path):
    openpyxl.Workbook().save(workbook_path)  # create workbook if non in folder
    print("New `pictures_import.xlsx` workbook created")



process_folders(workbook_path, base_folder="X:/AUR/11.2024/25.11.2024/import poze/poze ERP 25-11.zip")  # it unzipps automatically if needed
